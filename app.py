from flask import Flask, render_template, request, jsonify, send_file, make_response
import pymysql
import os
from werkzeug.utils import secure_filename
from xhtml2pdf import pisa
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)

def get_db_connection():
    connection = pymysql.connect(
        host='localhost',
        port=8111,
        user='root',
        password='',
        database='db_sisko',
        cursorclass=pymysql.cursors.DictCursor
    )
    return connection

@app.route('/')
def home():
    return render_template('index.html')

# ---- API LOGIN ----
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM users WHERE username = %s AND password = %s",
            (username, password)
        )
        user = cursor.fetchone()
        conn.close()

        if user:
            return jsonify({
                "success": True,
                "role": user['role'],
                "nisn": user['nisn'],
                "nama": user['nama_lengkap'],
                "username": user['username']
            })
        else:
            return jsonify({"success": False, "message": "Username atau password salah!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: AMBIL SEMUA DATA SISWA ----
@app.route('/api/siswa', methods=['GET'])
def get_siswa():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM data_siswa ORDER BY nama ASC")
        siswa = cursor.fetchall()
        conn.close()

        for s in siswa:
            if s['tgl_lahir']:
                s['tgl_lahir'] = s['tgl_lahir'].strftime('%Y-%m-%d')
            if s['timestamp']:
                s['timestamp'] = s['timestamp'].strftime('%d/%m/%Y %H:%M:%S')

        return jsonify({"success": True, "data": siswa})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: TAMBAH SISWA BARU ----
@app.route('/api/siswa', methods=['POST'])
def add_siswa():
    data = request.get_json()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id FROM data_siswa WHERE nisn = %s", (data.get('nisn'),))
        if cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "message": "NISN sudah terdaftar!"})

        cursor.execute("""
            INSERT INTO data_siswa
            (nama, nisn, jk, tgl_lahir, nama_ayah, nama_ibu, no_hp, kelas, jurusan, alamat, kode_pos)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data.get('nama'), data.get('nisn'), data.get('jk'),
            data.get('tglLahir') or None, data.get('namaAyah'), data.get('namaIbu'),
            data.get('noHp'), data.get('kelas'), data.get('jurusan'),
            data.get('alamat'), data.get('kodePos')
        ))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Siswa berhasil ditambahkan!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: EDIT DATA SISWA ----
@app.route('/api/siswa/<nisn>', methods=['PUT'])
def update_siswa(nisn):
    data = request.get_json()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE data_siswa SET
                nama = %s, jk = %s, tgl_lahir = %s, nama_ayah = %s,
                nama_ibu = %s, no_hp = %s, kelas = %s, jurusan = %s,
                alamat = %s, kode_pos = %s
            WHERE nisn = %s
        """, (
            data.get('nama'), data.get('jk'), data.get('tglLahir') or None,
            data.get('namaAyah'), data.get('namaIbu'), data.get('noHp'),
            data.get('kelas'), data.get('jurusan'), data.get('alamat'),
            data.get('kodePos'), nisn
        ))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Siswa berhasil diperbarui!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: HAPUS DATA SISWA ----
@app.route('/api/siswa/<nisn>', methods=['DELETE'])
def delete_siswa(nisn):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM data_siswa WHERE nisn = %s", (nisn,))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Siswa berhasil dihapus!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
# ---- API: UPLOAD FOTO SISWA ----
UPLOAD_FOLDER = os.path.join('static', 'uploads')

@app.route('/api/siswa/<nisn>/foto', methods=['POST'])
def upload_foto(nisn):
    try:
        if 'foto' not in request.files:
            return jsonify({"success": False, "message": "Tidak ada file yang dikirim!"})

        file = request.files['foto']
        if file.filename == '':
            return jsonify({"success": False, "message": "Nama file kosong!"})

        # Pastikan folder uploads ada, kalau belum, buat otomatis
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        # Amankan nama file & buat nama unik berdasarkan NISN
        ext = file.filename.rsplit('.', 1)[-1].lower()
        filename = secure_filename(f"{nisn}.{ext}")
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        # Simpan path relatif ke database (supaya bisa diakses lewat browser)
        foto_url = f"/static/uploads/{filename}"

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE data_siswa SET foto_url = %s WHERE nisn = %s", (foto_url, nisn))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Foto berhasil diupload!", "foto_url": foto_url})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: EXPORT PDF DATA SISWA ----
@app.route('/api/export/siswa/pdf', methods=['GET'])
def export_siswa_pdf():
    try:
        kelas = request.args.get('kelas', 'all')
        jurusan = request.args.get('jurusan', 'all')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil konfigurasi kop & TTD dulu
        cursor.execute("SELECT * FROM config")
        config_rows = cursor.fetchall()
        config = {row['key']: row['value'] for row in config_rows}

        # Ambil data siswa sesuai filter
        query = "SELECT * FROM data_siswa WHERE 1=1"
        params = []

        if kelas != 'all':
            query += " AND kelas = %s"
            params.append(kelas)
        if jurusan != 'all':
            query += " AND jurusan = %s"
            params.append(jurusan)

        query += " ORDER BY nama ASC"
        cursor.execute(query, params)
        siswa = cursor.fetchall()
        conn.close()

        # Format tanggal lahir supaya rapi tampil di PDF
        for s in siswa:
            if s['tgl_lahir']:
                s['tgl_lahir'] = s['tgl_lahir'].strftime('%d-%m-%Y')

        # Bangun isi tabel HTML baris per baris
        rows_html = ""
        if not siswa:
            rows_html = '<tr><td colspan="9" style="text-align:center; padding:20px;">Data tidak tersedia</td></tr>'
        else:
            for i, s in enumerate(siswa, start=1):
                rows_html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{s['nama']}</td>
                    <td>{s['nisn']}</td>
                    <td>{s['jk'] or '-'}</td>
                    <td>{s['tgl_lahir'] or '-'}</td>
                    <td>{s['kelas'] or '-'}</td>
                    <td>{s['jurusan'] or '-'}</td>
                    <td>{s['no_hp'] or '-'}</td>
                    <td>{s['alamat'] or '-'}</td>
                </tr>
                """

        subtitle = f"KELAS {kelas}" if kelas != 'all' else "SEMUA SISWA"

        kop1 = config.get('kop_instansi_1', 'PEMERINTAH PROVINSI')
        kop2 = config.get('kop_instansi_2', 'DINAS PENDIDIKAN')
        kop3 = config.get('kop_instansi_3', 'NAMA SEKOLAH')
        kop_alamat = config.get('kop_alamat', 'Alamat Sekolah')
        ttd_kota = config.get('ttd_kota', 'Kota')
        ttd_jabatan = config.get('ttd_jabatan', 'Kepala Sekolah')
        ttd_nama = config.get('ttd_nama', '..................')
        ttd_nip = config.get('ttd_nip', '..................')

        html_content = f"""
        <html>
        <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Helvetica, sans-serif; font-size: 10px; }}
            .kop {{ text-align: center; border-bottom: 3px double #000; padding-bottom: 8px; margin-bottom: 15px; }}
            .kop h3 {{ margin: 2px 0; font-size: 14px; }}
            .kop p {{ margin: 2px 0; font-size: 10px; font-style: italic; }}
            h2 {{ text-align: center; margin-bottom: 2px; text-decoration: underline; }}
            h4 {{ text-align: center; margin-top: 0; font-weight: normal; color: #555; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #333; padding: 5px; text-align: left; }}
            th {{ background-color: #2b2560; color: white; }}
            .ttd {{ margin-top: 30px; width: 100%; }}
            .ttd td {{ border: none; padding: 3px; }}
        </style>
        </head>
        <body>
            <div class="kop">
                <h3>{kop1}</h3>
                <h3>{kop2}</h3>
                <h2 style="font-size:18px; text-decoration:none;">{kop3}</h2>
                <p>{kop_alamat}</p>
            </div>

            <h2>LAPORAN DATA SISWA</h2>
            <h4>{subtitle}</h4>
            <table>
                <thead>
                    <tr>
                        <th>No</th><th>Nama</th><th>NISN</th><th>L/P</th><th>Tgl Lahir</th>
                        <th>Kelas</th><th>Jurusan</th><th>No HP</th><th>Alamat</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>

            <table class="ttd">
                <tr>
                    <td width="70%"></td>
                    <td width="30%" style="text-align:center;">
                        <p>{ttd_kota}, {datetime.now().strftime('%d-%m-%Y')}</p>
                        <p><b>{ttd_jabatan}</b></p>
                        <br><br><br>
                        <p><b><u>{ttd_nama}</u></b></p>
                        <p>NIP. {ttd_nip}</p>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """

        pdf_buffer = BytesIO()
        pisa.CreatePDF(html_content, dest=pdf_buffer)
        pdf_buffer.seek(0)

        response = make_response(pdf_buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=laporan_data_siswa.pdf'
        return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    
# ---- API: EXPORT EXCEL DATA SISWA ----
@app.route('/api/export/siswa/excel', methods=['GET'])
def export_siswa_excel():
    try:
        kelas = request.args.get('kelas', 'all')
        jurusan = request.args.get('jurusan', 'all')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil konfigurasi kop & TTD
        cursor.execute("SELECT * FROM config")
        config_rows = cursor.fetchall()
        config = {row['key']: row['value'] for row in config_rows}

        query = "SELECT * FROM data_siswa WHERE 1=1"
        params = []
        if kelas != 'all':
            query += " AND kelas = %s"
            params.append(kelas)
        if jurusan != 'all':
            query += " AND jurusan = %s"
            params.append(jurusan)
        query += " ORDER BY nama ASC"

        cursor.execute(query, params)
        siswa = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Siswa"

        headers = ["No", "Nama", "NISN", "L/P", "Tgl Lahir", "Nama Ayah", "Nama Ibu", "No HP", "Kelas", "Jurusan", "Alamat", "Kode Pos"]
        jumlah_kolom = len(headers)

        # ---- BAGIAN KOP SURAT ----
        ws.append([config.get('kop_instansi_1', '')])
        ws.append([config.get('kop_instansi_2', '')])
        ws.append([config.get('kop_instansi_3', '')])
        ws.append([config.get('kop_alamat', '')])
        ws.append([])  # baris kosong sebagai jarak

        for row_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=1)
            cell.font = Font(bold=(row_idx <= 3), size=(13 if row_idx == 3 else 11))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=jumlah_kolom)
            cell.alignment = cell.alignment.copy(horizontal='center')

        # ---- JUDUL LAPORAN ----
        judul_row = ws.max_row + 1
        subtitle = f"KELAS {kelas}" if kelas != 'all' else "SEMUA SISWA"
        ws.append(["LAPORAN DATA SISWA"])
        ws.append([subtitle])
        ws.append([])

        for r in [judul_row, judul_row + 1]:
            cell = ws.cell(row=r, column=1)
            cell.font = Font(bold=(r == judul_row), size=12)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=jumlah_kolom)
            cell.alignment = cell.alignment.copy(horizontal='center')

        # ---- HEADER TABEL ----
        header_row = ws.max_row + 1
        ws.append(headers)
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # ---- ISI DATA ----
        for i, s in enumerate(siswa, start=1):
            tgl_lahir = s['tgl_lahir'].strftime('%d-%m-%Y') if s['tgl_lahir'] else ''
            ws.append([
                i, s['nama'], s['nisn'], s['jk'], tgl_lahir,
                s['nama_ayah'], s['nama_ibu'], s['no_hp'],
                s['kelas'], s['jurusan'], s['alamat'], s['kode_pos']
            ])

        # ---- BLOK TANDA TANGAN ----
        ws.append([])
        ws.append([])
        ttd_kota = config.get('ttd_kota', 'Kota')
        ttd_jabatan = config.get('ttd_jabatan', 'Kepala Sekolah')
        ttd_nama = config.get('ttd_nama', '..................')
        ttd_nip = config.get('ttd_nip', '..................')
        tanggal_hari_ini = datetime.now().strftime('%d-%m-%Y')

        baris_ttd = [
            f"{ttd_kota}, {tanggal_hari_ini}",
            ttd_jabatan,
            "", "", "",
            ttd_nama,
            f"NIP. {ttd_nip}"
        ]
        for teks in baris_ttd:
            r = ws.max_row + 1
            ws.cell(row=r, column=jumlah_kolom - 2, value=teks)

        # Auto-lebar kolom biar rapi
        for col_cells in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col_cells:
                if cell.value and not isinstance(cell, type(ws.cell(row=1, column=1))) is False:
                    pass
                try:
                    col_letter = cell.column_letter
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                except AttributeError:
                    continue
            if col_letter:
                ws.column_dimensions[col_letter].width = max_length + 3

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = make_response(excel_buffer.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=laporan_data_siswa.xlsx'
        return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: EXPORT PDF DATA GURU ----
@app.route('/api/export/guru/pdf', methods=['GET'])
def export_guru_pdf():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil konfigurasi kop & TTD
        cursor.execute("SELECT * FROM config")
        config_rows = cursor.fetchall()
        config = {row['key']: row['value'] for row in config_rows}

        cursor.execute("SELECT * FROM data_guru ORDER BY nama ASC")
        guru = cursor.fetchall()
        conn.close()

        rows_html = ""
        if not guru:
            rows_html = '<tr><td colspan="6" style="text-align:center; padding:20px;">Data tidak tersedia</td></tr>'
        else:
            for i, g in enumerate(guru, start=1):
                rows_html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{g['nip']}</td>
                    <td>{g['nama']}</td>
                    <td>{g['mapel'] or '-'}</td>
                    <td>{g['kelas_ajar'] or '-'}</td>
                    <td>{g['jurusan_ajar'] or '-'}</td>
                </tr>
                """

        kop1 = config.get('kop_instansi_1', 'PEMERINTAH PROVINSI')
        kop2 = config.get('kop_instansi_2', 'DINAS PENDIDIKAN')
        kop3 = config.get('kop_instansi_3', 'NAMA SEKOLAH')
        kop_alamat = config.get('kop_alamat', 'Alamat Sekolah')
        ttd_kota = config.get('ttd_kota', 'Kota')
        ttd_jabatan = config.get('ttd_jabatan', 'Kepala Sekolah')
        ttd_nama = config.get('ttd_nama', '..................')
        ttd_nip = config.get('ttd_nip', '..................')

        html_content = f"""
        <html>
        <head>
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Helvetica, sans-serif; font-size: 11px; }}
            .kop {{ text-align: center; border-bottom: 3px double #000; padding-bottom: 8px; margin-bottom: 15px; }}
            .kop h3 {{ margin: 2px 0; font-size: 14px; }}
            .kop p {{ margin: 2px 0; font-size: 10px; font-style: italic; }}
            h2 {{ text-align: center; margin-bottom: 2px; text-decoration: underline; }}
            h4 {{ text-align: center; margin-top: 0; font-weight: normal; color: #555; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #333; padding: 6px; text-align: left; }}
            th {{ background-color: #2b2560; color: white; }}
            .ttd {{ margin-top: 30px; width: 100%; }}
            .ttd td {{ border: none; padding: 3px; }}
        </style>
        </head>
        <body>
            <div class="kop">
                <h3>{kop1}</h3>
                <h3>{kop2}</h3>
                <h2 style="font-size:18px; text-decoration:none;">{kop3}</h2>
                <p>{kop_alamat}</p>
            </div>

            <h2>LAPORAN DATA GURU</h2>
            <h4>DATA PENGAJAR AKTIF</h4>
            <table>
                <thead>
                    <tr><th>No</th><th>NIP</th><th>Nama Guru</th><th>Mata Pelajaran</th><th>Kelas Ajar</th><th>Jurusan Ajar</th></tr>
                </thead>
                <tbody>{rows_html}</tbody>
            </table>

            <table class="ttd">
                <tr>
                    <td width="70%"></td>
                    <td width="30%" style="text-align:center;">
                        <p>{ttd_kota}, {datetime.now().strftime('%d-%m-%Y')}</p>
                        <p><b>{ttd_jabatan}</b></p>
                        <br><br><br>
                        <p><b><u>{ttd_nama}</u></b></p>
                        <p>NIP. {ttd_nip}</p>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """

        pdf_buffer = BytesIO()
        pisa.CreatePDF(html_content, dest=pdf_buffer)
        pdf_buffer.seek(0)

        response = make_response(pdf_buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=laporan_data_guru.pdf'
        return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: EXPORT EXCEL DATA GURU ----
@app.route('/api/export/guru/excel', methods=['GET'])
def export_guru_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM config")
        config_rows = cursor.fetchall()
        config = {row['key']: row['value'] for row in config_rows}

        cursor.execute("SELECT * FROM data_guru ORDER BY nama ASC")
        guru = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Guru"

        headers = ["No", "NIP", "Nama Guru", "Mata Pelajaran", "Kelas Ajar", "Jurusan Ajar"]
        jumlah_kolom = len(headers)

        # ---- BAGIAN KOP SURAT ----
        ws.append([config.get('kop_instansi_1', '')])
        ws.append([config.get('kop_instansi_2', '')])
        ws.append([config.get('kop_instansi_3', '')])
        ws.append([config.get('kop_alamat', '')])
        ws.append([])

        for row_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=1)
            cell.font = Font(bold=(row_idx <= 3), size=(13 if row_idx == 3 else 11))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=jumlah_kolom)
            cell.alignment = cell.alignment.copy(horizontal='center')

        # ---- JUDUL LAPORAN ----
        judul_row = ws.max_row + 1
        ws.append(["LAPORAN DATA GURU"])
        ws.append(["DATA PENGAJAR AKTIF"])
        ws.append([])

        for r in [judul_row, judul_row + 1]:
            cell = ws.cell(row=r, column=1)
            cell.font = Font(bold=(r == judul_row), size=12)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=jumlah_kolom)
            cell.alignment = cell.alignment.copy(horizontal='center')

        # ---- HEADER TABEL ----
        header_row = ws.max_row + 1
        ws.append(headers)
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # ---- ISI DATA ----
        for i, g in enumerate(guru, start=1):
            ws.append([i, g['nip'], g['nama'], g['mapel'], g['kelas_ajar'], g['jurusan_ajar']])

        # ---- BLOK TANDA TANGAN ----
        ws.append([])
        ws.append([])
        ttd_kota = config.get('ttd_kota', 'Kota')
        ttd_jabatan = config.get('ttd_jabatan', 'Kepala Sekolah')
        ttd_nama = config.get('ttd_nama', '..................')
        ttd_nip = config.get('ttd_nip', '..................')
        tanggal_hari_ini = datetime.now().strftime('%d-%m-%Y')

        baris_ttd = [
            f"{ttd_kota}, {tanggal_hari_ini}",
            ttd_jabatan,
            "", "", "",
            ttd_nama,
            f"NIP. {ttd_nip}"
        ]
        for teks in baris_ttd:
            r = ws.max_row + 1
            ws.cell(row=r, column=jumlah_kolom - 2, value=teks)

        for col_cells in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col_cells:
                try:
                    col_letter = cell.column_letter
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                except AttributeError:
                    continue
            if col_letter:
                ws.column_dimensions[col_letter].width = max_length + 3

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = make_response(excel_buffer.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=laporan_data_guru.xlsx'
        return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: AMBIL DATA GURU (untuk preview) ----
@app.route('/api/guru', methods=['GET'])
def get_guru():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM data_guru ORDER BY nama ASC")
        guru = cursor.fetchall()
        conn.close()
        return jsonify({"success": True, "data": guru})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: TAMBAH GURU BARU ----
@app.route('/api/guru', methods=['POST'])
def add_guru():
    data = request.get_json()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id FROM data_guru WHERE nip = %s", (data.get('nip'),))
        if cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "message": "NIP Guru sudah terdaftar!"})

        cursor.execute("""
            INSERT INTO data_guru (nip, nama, mapel, kelas_ajar, jurusan_ajar)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            data.get('nip'), data.get('nama'), data.get('mapel'),
            data.get('kelasAjar'), data.get('jurusanAjar')
        ))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Guru berhasil disimpan!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


# ---- API: EDIT DATA GURU ----
@app.route('/api/guru/<nip>', methods=['PUT'])
def update_guru(nip):
    data = request.get_json()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE data_guru SET nama = %s, mapel = %s, kelas_ajar = %s, jurusan_ajar = %s
            WHERE nip = %s
        """, (
            data.get('nama'), data.get('mapel'),
            data.get('kelasAjar'), data.get('jurusanAjar'), nip
        ))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Guru berhasil diperbarui!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


# ---- API: HAPUS DATA GURU ----
@app.route('/api/guru/<nip>', methods=['DELETE'])
def delete_guru(nip):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM data_guru WHERE nip = %s", (nip,))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Data Guru berhasil dihapus!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ---- API: AMBIL KONFIGURASI KOP & TTD ----
@app.route('/api/config', methods=['GET'])
def get_config():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM config")
        rows = cursor.fetchall()
        conn.close()

        # Ubah dari format [{key: 'x', value: 'y'}, ...] jadi {x: 'y', ...} biar gampang dipakai
        config_map = {row['key']: row['value'] for row in rows}

        return jsonify({"success": True, "data": config_map})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


# ---- API: SIMPAN KONFIGURASI KOP & TTD ----
@app.route('/api/config', methods=['POST'])
def save_config():
    data = request.get_json()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        keys_to_save = [
            'kop_instansi_1', 'kop_instansi_2', 'kop_instansi_3', 'kop_alamat',
            'ttd_kota', 'ttd_jabatan', 'ttd_nama', 'ttd_nip'
        ]

        for key in keys_to_save:
            value = data.get(key, '')
            cursor.execute("""
                INSERT INTO config (`key`, value) VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE value = %s
            """, (key, value, value))

        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Konfigurasi berhasil diperbarui!"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
                            
if __name__ == '__main__':
    app.run(debug=True, port=5000, use_reloader=False)